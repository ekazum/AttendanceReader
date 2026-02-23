"""
AttendanceReader - Malam Sachar Plus PDF to Excel converter.

Reads an Israeli "Malam Sachar Plus" monthly attendance PDF and converts it
into a structured Excel (.xlsx) file.

Build command (PyInstaller):
    pyinstaller --onefile --windowed --name AttendanceReader attendance_reader.py
"""

import os
import re
import threading
import tkinter as tk
from tkinter import filedialog, ttk

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pdfplumber

# ---------------------------------------------------------------------------
# Hebrew day-letter -> full day name
# ---------------------------------------------------------------------------
HEBREW_DAYS = {
    "א": "ראשון",   # Sunday
    "ב": "שני",     # Monday
    "ג": "שלישי",   # Tuesday
    "ד": "רביעי",   # Wednesday
    "ה": "חמישי",   # Thursday
    "ו": "שישי",    # Friday
    "ש": "שבת",     # Saturday
}

DATE_RE = re.compile(r"^\d{2}/\d{2}$")
TIME_RE = re.compile(r"^\d{2}:\d{2}$")


def _is_date_row(texts: list) -> bool:
    """Return True if *texts* contains 3 or more DD/MM date tokens."""
    return sum(1 for t in texts if DATE_RE.match(t)) >= 3


# ---------------------------------------------------------------------------
# PDF extraction helpers
# ---------------------------------------------------------------------------

def _x_center(word: dict) -> float:
    return (word["x0"] + word["x1"]) / 2.0


def _group_rows(words: list, y_tolerance: float = 4.0) -> list:
    """Group word dicts by approximate y-position and return sorted list of
    (y, [words]) tuples ordered top-to-bottom."""
    buckets: dict[float, list] = {}
    for w in words:
        y = w["top"]
        matched = next(
            (k for k in buckets if abs(k - y) <= y_tolerance), None
        )
        if matched is None:
            buckets[y] = []
            matched = y
        buckets[matched].append(w)
    return sorted(buckets.items())


def _find_closest(lookup: dict, x: float, tolerance: float) -> str:
    """Return the value in *lookup* whose key is closest to *x*, or '' if
    no key is within *tolerance*."""
    if not lookup:
        return ""
    closest = min(lookup.keys(), key=lambda k: abs(k - x))
    return lookup[closest] if abs(closest - x) <= tolerance else ""


def _build_lookup(row_words: list, pattern=None) -> dict:
    """Build {x_center: text} dict from a row, optionally filtering by regex."""
    result = {}
    for w in row_words:
        if pattern is None or pattern.match(w["text"]):
            result[_x_center(w)] = w["text"]
    return result


def _build_data_lookup(
    row_words: list,
    data_x_min: float,
    data_x_max: float,
    tolerance: float,
    pattern=None,
) -> dict:
    """Build {x_center: text} for words whose x-center falls within the data
    column range [data_x_min - tolerance, data_x_max + tolerance].

    This excludes row-label words that sit outside the column grid.
    If *pattern* is provided, only words matching it are included (for time rows).
    """
    result = {}
    for w in row_words:
        cx = _x_center(w)
        if cx < data_x_min - tolerance or cx > data_x_max + tolerance:
            continue
        text = w["text"]
        if pattern is None or pattern.match(text):
            result[cx] = text
    return result


def _row_label(row_words: list) -> str:
    """Return a single string formed by joining all non-time, non-date words in
    the row.  Used to identify which semantic role a row plays (entry, exit,
    total, day-type, …)."""
    return " ".join(
        w["text"]
        for w in row_words
        if not TIME_RE.match(w["text"]) and not DATE_RE.match(w["text"])
    )


def _process_block(block_rows: list) -> list:
    """Process one horizontal grid block (date row + its associated data rows).

    Rows are classified by their Hebrew label text rather than by position
    index, so overtime rows and other extra rows are safely ignored.
    """
    date_row_words = None
    day_row_words = None
    entry_row_words = None
    exit_row_words = None
    total_row_words = None
    day_type_row_words = None

    for _y, row_words in block_rows:
        texts = [w["text"] for w in row_words]

        # Date row (3+ DD/MM tokens) – already the first row of the block but
        # handle it here for completeness.
        if _is_date_row(texts):
            if date_row_words is None:
                date_row_words = row_words
            continue

        # Day-letter row (3+ single Hebrew day letters)
        if sum(1 for t in texts if t in HEBREW_DAYS) >= 3:
            if day_row_words is None:
                day_row_words = row_words
            continue

        # Classify remaining rows by their Hebrew label
        label = _row_label(row_words)

        if "שעת כניסה" in label and entry_row_words is None:
            entry_row_words = row_words
        elif "שעת יציאה" in label and exit_row_words is None:
            exit_row_words = row_words
        elif "סה" in label and "נוכח" in label and total_row_words is None:
            total_row_words = row_words
        elif ("סוג יום" in label or "נוכחות" in label) and day_type_row_words is None:
            day_type_row_words = row_words

    if not date_row_words:
        return []

    # Collect date columns sorted left-to-right
    date_cols = sorted(
        [(_x_center(w), w["text"]) for w in date_row_words if DATE_RE.match(w["text"])],
        key=lambda p: p[0],
    )
    if not date_cols:
        return []

    # Adaptive column-matching tolerance: max(45% of avg inter-column spacing, 10 pt)
    xs = [x for x, _ in date_cols]
    data_x_min, data_x_max = xs[0], xs[-1]
    if len(date_cols) > 1:
        avg_spacing = (data_x_max - data_x_min) / (len(xs) - 1)
        tolerance = max(avg_spacing * 0.45, 10.0)
    else:
        tolerance = 15.0

    # Day lookup: only HEBREW_DAYS letters (excludes label words automatically)
    day_lookup = {
        _x_center(w): w["text"]
        for w in (day_row_words or [])
        if w["text"] in HEBREW_DAYS
    }

    entry_lookup = (
        _build_data_lookup(entry_row_words, data_x_min, data_x_max, tolerance, TIME_RE)
        if entry_row_words else {}
    )
    exit_lookup = (
        _build_data_lookup(exit_row_words, data_x_min, data_x_max, tolerance, TIME_RE)
        if exit_row_words else {}
    )
    total_lookup = (
        _build_data_lookup(total_row_words, data_x_min, data_x_max, tolerance, TIME_RE)
        if total_row_words else {}
    )
    # Day-type: no pattern filter – values are free Hebrew text (e.g. "חופשה")
    day_type_lookup = (
        _build_data_lookup(day_type_row_words, data_x_min, data_x_max, tolerance)
        if day_type_row_words else {}
    )

    records = []
    for x, date in date_cols:
        day_letter = _find_closest(day_lookup, x, tolerance)
        day_name = HEBREW_DAYS.get(day_letter, day_letter)
        entry = _find_closest(entry_lookup, x, tolerance)
        exit_time = _find_closest(exit_lookup, x, tolerance)
        total = _find_closest(total_lookup, x, tolerance)
        day_type = _find_closest(day_type_lookup, x, tolerance)

        records.append(
            {
                "date": date,
                "day": day_name,
                "entry": entry,
                "exit": exit_time,
                "total": total,
                "day_type": day_type,
            }
        )

    return records


def _process_page(page) -> list:
    """Extract attendance records from a single pdfplumber page object.

    A page may contain multiple horizontal grid blocks (e.g. days 1-15 on top,
    days 16-31 below).  Each block is identified by a date row (3+ DD/MM
    tokens) and processed independently.
    """
    words = page.extract_words(
        x_tolerance=3, y_tolerance=3, keep_blank_chars=False
    )
    if not words:
        return []

    sorted_rows = _group_rows(words)

    # Split rows into blocks; a new block begins whenever a date row is found.
    blocks: list = []
    current_block: list = []

    for y, row_words in sorted_rows:
        texts = [w["text"] for w in row_words]
        if _is_date_row(texts):
            if current_block:
                blocks.append(current_block)
            current_block = [(y, row_words)]
        elif current_block:
            current_block.append((y, row_words))

    if current_block:
        blocks.append(current_block)

    records = []
    for block in blocks:
        records.extend(_process_block(block))
    return records


def extract_attendance_data(pdf_path: str) -> list:
    """Open the PDF and return a list of attendance record dicts."""
    records = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            records.extend(_process_page(page))
    return records


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADERS = ["תאריך", "יום בשבוע", "שעת כניסה", "שעת יציאה", 'סה"כ שעות ליום', "סוג יום"]
COLUMN_WIDTHS = [12, 15, 14, 14, 18, 14]


def create_excel(records: list, output_path: str) -> None:
    """Write *records* to an Excel file at *output_path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "נוכחות"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Write data
    for row_idx, record in enumerate(records, 2):
        values = [
            record["date"],
            record["day"],
            record["entry"],
            record["exit"],
            record["total"],
            record.get("day_type", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = center

    # Column widths
    for col, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Usability helpers
    ws.auto_filter.ref = f"A1:F{len(records) + 1}"
    ws.freeze_panes = "A2"

    wb.save(output_path)


# ---------------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------------

class AttendanceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ממיר נוכחות – PDF לאקסל")
        self.geometry("540x210")
        self.resizable(False, False)
        self._configure_rtl()

        self.pdf_path = tk.StringVar()
        self._build_ui()

    def _configure_rtl(self):
        self.option_add("*font", "Arial 10")
        # Make the window appear right-to-left by default
        try:
            self.tk.call("encoding", "system", "utf-8")
        except tk.TclError:
            pass

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        # ── PDF selection row ──────────────────────────────────────────────
        row1 = tk.Frame(self)
        row1.pack(fill=tk.X, **pad)

        tk.Button(row1, text="בחר קובץ PDF", command=self._browse_pdf,
                  width=14).pack(side=tk.RIGHT)
        tk.Entry(row1, textvariable=self.pdf_path, width=42,
                 state="readonly").pack(side=tk.RIGHT, padx=(0, 6))

        # ── Convert button ─────────────────────────────────────────────────
        row2 = tk.Frame(self)
        row2.pack(fill=tk.X, **pad)

        self.convert_btn = tk.Button(
            row2, text="המר לאקסל  ▶",
            command=self._start_conversion,
            bg="#4CAF50", fg="white",
            font=("Arial", 11, "bold"),
            padx=16, pady=4,
        )
        self.convert_btn.pack(side=tk.RIGHT)

        # ── Status label ───────────────────────────────────────────────────
        row3 = tk.Frame(self)
        row3.pack(fill=tk.X, **pad)

        self.status_var = tk.StringVar(value="נא לבחור קובץ PDF ולחץ 'המר לאקסל'")
        tk.Label(row3, textvariable=self.status_var,
                 wraplength=520, justify="right",
                 anchor="e").pack(fill=tk.X)

        # ── Progress bar ───────────────────────────────────────────────────
        self.progress = ttk.Progressbar(self, mode="indeterminate")
        self.progress.pack(fill=tk.X, padx=10, pady=(0, 8))

    # ------------------------------------------------------------------
    # Event handlers
    # ------------------------------------------------------------------

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="בחר קובץ נוכחות PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if path:
            self.pdf_path.set(path)
            self.status_var.set(f"נבחר: {os.path.basename(path)}")

    def _start_conversion(self):
        pdf = self.pdf_path.get()
        if not pdf:
            self.status_var.set("שגיאה: לא נבחר קובץ PDF")
            return
        if not os.path.exists(pdf):
            self.status_var.set("שגיאה: הקובץ אינו קיים")
            return

        default_name = os.path.splitext(os.path.basename(pdf))[0] + ".xlsx"
        output = filedialog.asksaveasfilename(
            title="שמור קובץ אקסל",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_name,
        )
        if not output:
            return  # user cancelled

        self.convert_btn.config(state=tk.DISABLED)
        self.progress.start(12)
        self.status_var.set("ממיר, נא המתן…")

        threading.Thread(
            target=self._convert_worker, args=(pdf, output), daemon=True
        ).start()

    def _convert_worker(self, pdf_path: str, output_path: str):
        try:
            records = extract_attendance_data(pdf_path)
            if not records:
                self.after(0, self._on_error, "לא נמצאו נתוני נוכחות בקובץ")
                return
            create_excel(records, output_path)
            self.after(0, self._on_success, output_path, len(records))
        except (FileNotFoundError, PermissionError, OSError) as exc:
            self.after(0, self._on_error, f"שגיאת קובץ: {exc.strerror}")
        except Exception as exc:  # noqa: BLE001
            self.after(0, self._on_error, str(exc))

    def _on_success(self, output_path: str, count: int):
        self.progress.stop()
        self.convert_btn.config(state=tk.NORMAL)
        self.status_var.set(
            f"✔ הצלחה! {count} ימים יוצאו → {os.path.basename(output_path)}"
        )

    def _on_error(self, message: str):
        self.progress.stop()
        self.convert_btn.config(state=tk.NORMAL)
        self.status_var.set(f"✘ שגיאה: {message}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = AttendanceApp()
    app.mainloop()
