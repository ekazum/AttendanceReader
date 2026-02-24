"""
malam_saar_attendance.py – Malam Saar Hebrew attendance PDF → Excel converter.

Parses multi-month attendance PDFs produced by the Israeli "מלם שכר" (Malam Saar)
payroll system, extracting per-day records and writing a formatted Excel workbook.

Build command (PyInstaller – single .exe, no console window):
    pyinstaller --onefile --windowed --name MalamSaarAttendance \\
        --add-data "columns_config.json;." \\
        malam_saar_attendance.py

Place columns_config.json next to the .exe to override default column x-ranges.
"""

from __future__ import annotations

import json
import logging
import os
import re
import sys
import threading
import tkinter as tk
from datetime import date
from tkinter import filedialog, scrolledtext, ttk
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pdfplumber

# ---------------------------------------------------------------------------
# Logging – write to an in-memory list that the GUI drains
# ---------------------------------------------------------------------------

_log_messages: list[str] = []


def _log(msg: str) -> None:
    _log_messages.append(msg)
    logging.info(msg)


# ---------------------------------------------------------------------------
# RTL / Hebrew normalisation
# ---------------------------------------------------------------------------

_HEBREW_CHAR_RE = re.compile(r"[\u05D0-\u05EA]")


def normalize_hebrew_word(word: str) -> str:
    """Reverse *word* if it contains Hebrew characters; leave everything else unchanged.

    This reversal is required because pdfplumber extracts Hebrew text in visual
    (left-to-right) PDF order, which reverses the logical character sequence.
    Numbers, times (08:00), dates (12/02), and ASCII symbols are not reversed.
    """
    if _HEBREW_CHAR_RE.search(word):
        return word[::-1]
    return word


def normalize_row_words(words: list[dict]) -> list[dict]:
    """Apply normalize_hebrew_word to the 'text' field of each pdfplumber word dict.

    Call immediately after page.extract_words() before any other processing.
    """
    for w in words:
        w["text"] = normalize_hebrew_word(w["text"])
    return words


def normalize_line(line: str) -> str:
    """Reverse each Hebrew token in a whitespace-delimited line; leave others intact."""
    return " ".join(normalize_hebrew_word(tok) for tok in line.split())


# ---------------------------------------------------------------------------
# Configuration loader
# ---------------------------------------------------------------------------

DEFAULT_CONFIG: dict[str, list[float]] = {
    "date":           [795, 830],
    "day_of_week":    [782, 800],
    "shift_marker":   [750, 765],
    "entry_actual":   [704, 720],
    "exit_actual":    [677, 693],
    "total_present":  [652, 668],   # was [620, 637] — FIXED
    "entry_for_pay":  [620, 645],   # was [652, 668] — FIXED
    "exit_for_pay":   [591, 607],
    "total_for_pay":  [565, 581],
    "activity":       [485, 545],   # widened — FIXED (Bug 3)
    "standard_hours": [434, 492],
    "ot_100":         [398, 420],
    "ot_125":         [362, 382],   # was [255, 277] — FIXED
    "ot_150":         [327, 347],   # was [291, 313] — FIXED
    "ot_200":         [291, 311],   # was [327, 349] — FIXED
    "shift_87":       [251, 272],   # was [184, 210] — FIXED
    "shift_50":       [222, 242],
    "shift_20":       [184, 207],   # was [256, 282] — FIXED
    "deduction":      [110, 135],
}


def load_config() -> dict[str, list[float]]:
    """Load columns_config.json from next to the executable (or script).

    If the file is absent, the default config is written there and returned.
    Works both in development and when frozen by PyInstaller.
    """
    # Determine directory that contains the running exe / script
    exe_dir = os.path.dirname(
        sys.executable if getattr(sys, "frozen", False)
        else os.path.abspath(__file__)
    )
    user_config = os.path.join(exe_dir, "columns_config.json")

    # Bundled fallback path (PyInstaller _MEIPASS)
    bundled_config = os.path.join(
        getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__))),
        "columns_config.json",
    )

    if os.path.exists(user_config):
        config_path = user_config
    elif os.path.exists(bundled_config):
        config_path = bundled_config
    else:
        # Write the default and use it
        try:
            with open(user_config, "w", encoding="utf-8") as fh:
                json.dump(DEFAULT_CONFIG, fh, ensure_ascii=False, indent=2)
        except OSError:
            pass
        return dict(DEFAULT_CONFIG)

    with open(config_path, encoding="utf-8") as fh:
        return json.load(fh)


# ---------------------------------------------------------------------------
# Activity type mapping
# ---------------------------------------------------------------------------

ACTIVITY_MAP: dict[str, str] = {
    "עבודה":              "Work",
    "חופשה":              "Vacation",
    "מחלה":               "Sick",
    "אין דיווח נוכחות":  "No Report",
    "ללא תקן עבודה":      "No Standard",
    "בחירה":              "Personal Day",
    "מחלת בן זוג":        "Spouse Sick",
    "מחלת הורה":          "Parent Sick",
    "מחלה בהצהרה":        "Declared Sick",
    "השתלמות":            "Training",
    "חג":                 "Holiday",
    'חוה"מ':              "Intermediate Holiday",
    "ע' חג":              "Holiday Eve",
}

# Activity values for which all time/OT columns are set to None
_NULL_ACTIVITIES = {"אין דיווח נוכחות", "ללא תקן עבודה"}

# Regex patterns (applied *after* normalisation → visual/token order preserved,
# numbers appear BEFORE Hebrew labels in the extracted line)
_RE_SALARY_MONTH = re.compile(r"(\d{2}/\d{2}/\d{4})\s+שכר\s+בחודש")
_RE_EMPLOYEE_ID = re.compile(r"(\d{7,9})\s+זהות")
_RE_TAG_NUMBER = re.compile(r"(\d+)\s+תג:")
_RE_DATE_TOKEN = re.compile(r"^\d{2}/\d{2}$")
_RE_TIME_TOKEN = re.compile(r"^\d{2}:\d{2}(:\d{2})?$")
_RE_DATA_PAGE_MARKER = re.compile(r"חישוב:")


def _extract_employee_name(normalized_text: str) -> str:
    """Extract employee name from the normalised header line.

    After normalisation the relevant line looks like:
      'בשבוע העבודה ימי ויאצ'סלב פונדר שם מ.נ.0 32070758 זהות'
    The employee name is the sequence of consecutive Hebrew-only words
    immediately to the LEFT of the token 'שם' (i.e. preceding it in the string).
    We split the line on 'שם', take everything before it, then walk backwards
    collecting Hebrew words until we hit a non-Hebrew token.
    """
    for line in normalized_text.splitlines():
        if "שם" in line and "זהות" in line:
            before_shem = line.split("שם")[0]
            tokens = before_shem.strip().split()
            name_parts: list[str] = []
            for token in reversed(tokens):
                if _HEBREW_CHAR_RE.search(token):
                    name_parts.insert(0, token)
                else:
                    break
            if name_parts:
                return " ".join(name_parts)
    return ""


# ---------------------------------------------------------------------------
# Column assignment helpers
# ---------------------------------------------------------------------------

def _assign_column(x0: float, config: dict[str, list[float]]) -> str | None:
    """Return the column name whose x-range contains *x0*, or None."""
    for col, (lo, hi) in config.items():
        if lo <= x0 <= hi:
            return col
    return None


def _parse_standard_hours(token: str) -> tuple[str | None, str | None]:
    """Handle the 'standard_hours' merged-token case.

    Two adjacent PDF cells are sometimes concatenated by pdfplumber, e.g.
    '08:0008:00' (10 chars: two HH:MM tokens back-to-back).
    The spec also notes the length-11 / token[5]==':' variant.
    In both cases we take the first 5 characters as the standard_hours value
    and discard the rest (redundant duplicate).

    Returns (standard_hours_value, None).
    """
    # Merged double-time: length 10 ("08:0008:00") or 11 with separator
    if len(token) >= 10 and _RE_TIME_TOKEN.match(token[:5]):
        return token[:5], None
    # Normal time
    if _RE_TIME_TOKEN.match(token):
        return token, None
    return token, None


# ---------------------------------------------------------------------------
# Date reconstruction
# ---------------------------------------------------------------------------

def _reconstruct_date(dd: int, mm: int, salary_month_str: str) -> date | None:
    """Reconstruct a full date from a DD/MM token and the salary_month header.

    salary_month_str is formatted '01/MM/YYYY'.
    """
    try:
        parts = salary_month_str.split("/")
        salary_year = int(parts[2])
        salary_mm = int(parts[1])
    except (IndexError, ValueError):
        return None

    try:
        if mm <= salary_mm:
            return date(salary_year, mm, dd)
        # row_month > salary_month_num → date belongs to the prior year
        return date(salary_year - 1, mm, dd)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Header extraction from a single data page
# ---------------------------------------------------------------------------

def _extract_headers(page_text: str) -> dict[str, str]:
    """Extract salary_month, employee_name, employee_id, tag_number from page text.

    *page_text* must already be normalised (each line passed through normalize_line).
    Numbers appear BEFORE their Hebrew labels after normalisation.
    """
    headers: dict[str, str] = {}

    m = _RE_SALARY_MONTH.search(page_text)
    if m:
        headers["salary_month"] = m.group(1)

    m = _RE_EMPLOYEE_ID.search(page_text)
    if m:
        headers["employee_id"] = m.group(1)

    m = _RE_TAG_NUMBER.search(page_text)
    if m:
        headers["tag_number"] = m.group(1)

    name = _extract_employee_name(page_text)
    if name:
        headers["employee_name"] = name

    return headers


# ---------------------------------------------------------------------------
# Page-level row parsing
# ---------------------------------------------------------------------------

def _group_by_y(words: list[dict], tolerance: float = 4.0) -> dict[int, list[dict]]:
    """Group words into rows by rounding their y-coordinate to the nearest bucket."""
    rows: dict[int, list[dict]] = {}
    for w in words:
        y = w.get("top", 0.0)
        bucket = int(round(y / tolerance) * tolerance)
        rows.setdefault(bucket, []).append(w)
    return rows


def _parse_data_rows(
    words: list[dict],
    config: dict[str, list[float]],
    salary_month: str,
    page_num: int,
) -> list[dict[str, Any]]:
    """Parse daily attendance rows from a list of normalised word dicts.

    Returns a list of record dicts.
    """
    date_range = config.get("date", [795, 830])
    grouped = _group_by_y(words)

    records: list[dict[str, Any]] = []

    for _bucket, row_words in sorted(grouped.items()):
        # Check whether this row has a date token in the date column
        date_token = None
        for w in row_words:
            if _RE_DATE_TOKEN.match(w["text"]) and date_range[0] <= w["x0"] <= date_range[1]:
                date_token = w["text"]
                break
        if not date_token:
            continue

        # Parse DD/MM
        try:
            dd, mm = int(date_token[:2]), int(date_token[3:5])
        except ValueError:
            continue

        full_date = _reconstruct_date(dd, mm, salary_month)

        # Build column→value map for this row
        col_values: dict[str, Any] = {}
        activity_words: list[tuple[float, str]] = []  # (x0, text)
        for w in row_words:
            col = _assign_column(w["x0"], config)
            if col is None:
                continue
            text = w["text"]

            if col == "standard_hours":
                std, _ = _parse_standard_hours(text)
                col_values[col] = std
            elif col == "shift_marker":
                col_values["shift_premium"] = (text == "*")
            elif col == "activity":
                activity_words.append((w["x0"], text))
            else:
                col_values[col] = text

        # Sort activity words by descending x (rightmost first = correct RTL order)
        activity_words.sort(key=lambda t: t[0], reverse=True)
        col_values["activity"] = " ".join(text for _, text in activity_words)

        activity_raw = col_values.get("activity", "").strip()
        activity_en = ACTIVITY_MAP.get(activity_raw, activity_raw)

        _log(
            f"Page {page_num} Row {date_token}: "
            f"{activity_raw} -> {activity_en}"
        )

        # Null out time/OT for certain activity types
        null_time = activity_raw in _NULL_ACTIVITIES

        def _t(key: str) -> str | None:
            return None if null_time else col_values.get(key)

        salary_month_display = (
            salary_month[3:] if salary_month else ""  # 'MM/YYYY'
        )

        record: dict[str, Any] = {
            "salary_month":          salary_month_display,
            "full_date":             full_date,
            "day_of_week":           col_values.get("day_of_week"),
            "day_type":              activity_en,
            "shift_premium":         col_values.get("shift_premium", False),
            "entry_actual":          _t("entry_actual"),
            "exit_actual":           _t("exit_actual"),
            "total_present_hours":   _t("total_present"),
            "entry_for_pay":         _t("entry_for_pay"),
            "exit_for_pay":          _t("exit_for_pay"),
            "total_for_pay_hours":   _t("total_for_pay"),
            "standard_hours":        _t("standard_hours"),
            "ot_100":                _t("ot_100"),
            "ot_125":                _t("ot_125"),
            "ot_150":                _t("ot_150"),
            "ot_200":                _t("ot_200"),
            "shift_bonus_87":        _t("shift_87"),
            "shift_bonus_50":        _t("shift_50"),
            "shift_bonus_20":        _t("shift_20"),
            "deduction":             _t("deduction"),
        }
        records.append(record)

    return records


# ---------------------------------------------------------------------------
# Main PDF processing
# ---------------------------------------------------------------------------

def process_pdf(pdf_path: str) -> list[dict[str, Any]]:
    """Parse *pdf_path* and return a flat list of daily attendance record dicts.

    Config is reloaded on every call (never cached).
    """
    config = load_config()
    all_records: list[dict[str, Any]] = []

    # Per-employee header fields – carried forward across pages
    current_employee: dict[str, str] = {
        "employee_id":   "",
        "employee_name": "",
        "tag_number":    "",
        "salary_month":  "01/01/1900",
    }

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            try:
                raw_text = page.extract_text() or ""
                # Normalise each line for header regex matching
                normalised_text = "\n".join(
                    normalize_line(line) for line in raw_text.splitlines()
                )

                # Skip summary pages (those without the data-page marker)
                if not _RE_DATA_PAGE_MARKER.search(normalised_text):
                    _log(f"Page {page_num}: SUMMARY – skipped")
                    continue

                # Update employee header fields from this page
                headers = _extract_headers(normalised_text)
                if headers.get("employee_id"):
                    current_employee["employee_id"] = headers["employee_id"]
                if headers.get("employee_name"):
                    current_employee["employee_name"] = headers["employee_name"]
                if headers.get("tag_number"):
                    current_employee["tag_number"] = headers["tag_number"]
                if headers.get("salary_month"):
                    current_employee["salary_month"] = headers["salary_month"]

                # Extract and normalise word tokens
                words = page.extract_words(x_tolerance=3, y_tolerance=3)
                normalize_row_words(words)

                page_records = _parse_data_rows(
                    words,
                    config,
                    current_employee["salary_month"],
                    page_num,
                )

                # Attach employee header fields to every record
                for rec in page_records:
                    rec["employee_id"] = current_employee["employee_id"]
                    rec["employee_name"] = current_employee["employee_name"]
                    rec["tag_number"] = current_employee["tag_number"]

                all_records.extend(page_records)
                _log(
                    f"Page {page_num}: DATA – "
                    f"{len(page_records)} rows, "
                    f"month={current_employee['salary_month']}"
                )

            except (OSError, ValueError, AttributeError, KeyError,
                    TypeError, RuntimeError) as exc:  # noqa: BLE001
                _log(f"Page {page_num}: ERROR – {type(exc).__name__}: {exc}")

    return all_records


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

# Sheet 1 column definitions: (header_label, record_key)
_DAILY_COLUMNS: list[tuple[str, str]] = [
    ("Employee ID",           "employee_id"),
    ("Employee Name",         "employee_name"),
    ("Tag Number",            "tag_number"),
    ("Salary Month",          "salary_month"),
    ("Date",                  "full_date"),
    ("Day of Week",           "day_of_week"),
    ("Day Type",              "day_type"),
    ("Shift Premium",         "shift_premium"),
    ("Entry (Actual)",        "entry_actual"),
    ("Exit (Actual)",         "exit_actual"),
    ("Total Present Hours",   "total_present_hours"),
    ("Entry (For Pay)",       "entry_for_pay"),
    ("Exit (For Pay)",        "exit_for_pay"),
    ("Total For Pay Hours",   "total_for_pay_hours"),
    ("Standard Hours",        "standard_hours"),
    ("OT 100%",               "ot_100"),
    ("OT 125%",               "ot_125"),
    ("OT 150%",               "ot_150"),
    ("OT 200%",               "ot_200"),
    ("Shift Bonus 87",        "shift_bonus_87"),
    ("Shift Bonus 50",        "shift_bonus_50"),
    ("Shift Bonus 20",        "shift_bonus_20"),
    ("Deduction",             "deduction"),
]

# Sheet 2 column positions in Sheet 1 (1-based, for SUMIF formulas)
# These are the column letters for Sheet 1 columns used in Sheet 2 formulas:
#   D = salary_month, E = full_date, K = total_present_hours,
#   N = total_for_pay_hours, P = ot_100, Q = ot_125, R = ot_150, S = ot_200
# Day type is column G; the day-type-specific counts use SUMIF on G.
_S1_SALARY_MONTH_COL   = "D"
_S1_DAY_TYPE_COL       = "G"
_S1_TOTAL_PRESENT_COL  = "K"
_S1_TOTAL_FOR_PAY_COL  = "N"
_S1_OT100_COL          = "P"
_S1_OT125_COL          = "Q"
_S1_OT150_COL          = "R"
_S1_OT200_COL          = "S"

_HEADER_FILL   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_HEADER_FONT   = Font(name="Arial", size=10, bold=True)
_DATA_FONT     = Font(name="Arial", size=10)
_ALT_FILL      = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_WHITE_FILL    = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT_ALIGN   = Alignment(horizontal="right", vertical="center")
_LEFT_ALIGN    = Alignment(horizontal="left", vertical="center")

# Columns whose data cells should be right-aligned
_RIGHT_ALIGN_KEYS = {
    "entry_actual", "exit_actual", "total_present_hours",
    "entry_for_pay", "exit_for_pay", "total_for_pay_hours",
    "standard_hours", "ot_100", "ot_125", "ot_150", "ot_200",
    "shift_bonus_87", "shift_bonus_50", "shift_bonus_20", "deduction",
}


def _auto_fit_column(ws, col_idx: int) -> None:
    """Set column width to max cell content length × 1.2, minimum 10."""
    max_len = 0
    col_letter = get_column_letter(col_idx)
    for cell in ws[col_letter]:
        try:
            cell_len = len(str(cell.value)) if cell.value is not None else 0
            max_len = max(max_len, cell_len)
        except (AttributeError, TypeError):
            pass
    ws.column_dimensions[col_letter].width = max(max_len * 1.2, 10)


def create_excel(records: list[dict[str, Any]], output_path: str) -> None:
    """Write *records* to a two-sheet Excel workbook at *output_path*."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Daily Attendance ─────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Daily Attendance"

    # Write header row
    for col_idx, (header, _key) in enumerate(_DAILY_COLUMNS, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # Write data rows
    for row_idx, record in enumerate(records, 2):
        fill = _ALT_FILL if (row_idx % 2 == 0) else _WHITE_FILL
        for col_idx, (_header, key) in enumerate(_DAILY_COLUMNS, 1):
            value = record.get(key)
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _DATA_FONT
            cell.fill = fill
            if key == "full_date" and isinstance(value, date):
                cell.number_format = "DD/MM/YYYY"
                cell.alignment = _RIGHT_ALIGN
            elif key in _RIGHT_ALIGN_KEYS:
                cell.alignment = _RIGHT_ALIGN
            else:
                cell.alignment = _LEFT_ALIGN

    # Freeze top row and auto-fit columns
    ws1.freeze_panes = "A2"
    for col_idx in range(1, len(_DAILY_COLUMNS) + 1):
        _auto_fit_column(ws1, col_idx)

    # ── Sheet 2: Monthly Summary ──────────────────────────────────────────
    ws2 = wb.create_sheet(title="Monthly Summary")

    summary_headers = [
        "Salary Month",
        "Work Days",
        "Vacation Days",
        "Sick Days",
        "No Report Days",
        "Total Present Hours",
        "Total For Pay Hours",
        "Total OT 100%",
        "Total OT 125%",
        "Total OT 150%",
        "Total OT 200%",
    ]

    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # Collect unique salary months in order of first appearance
    seen_months: list[str] = []
    for rec in records:
        sm = rec.get("salary_month", "")
        if sm and sm not in seen_months:
            seen_months.append(sm)

    total_rows = len(records) + 1  # data on Sheet 1 starts at row 2

    for sum_row, month in enumerate(seen_months, 2):
        # Column A: salary month string
        ws2.cell(row=sum_row, column=1, value=month).font = _DATA_FONT

        ref_col   = f"'Daily Attendance'!{_S1_SALARY_MONTH_COL}:{_S1_SALARY_MONTH_COL}"
        type_col  = f"'Daily Attendance'!{_S1_DAY_TYPE_COL}:{_S1_DAY_TYPE_COL}"
        pres_col  = f"'Daily Attendance'!{_S1_TOTAL_PRESENT_COL}:{_S1_TOTAL_PRESENT_COL}"
        pay_col   = f"'Daily Attendance'!{_S1_TOTAL_FOR_PAY_COL}:{_S1_TOTAL_FOR_PAY_COL}"
        ot100_col = f"'Daily Attendance'!{_S1_OT100_COL}:{_S1_OT100_COL}"
        ot125_col = f"'Daily Attendance'!{_S1_OT125_COL}:{_S1_OT125_COL}"
        ot150_col = f"'Daily Attendance'!{_S1_OT150_COL}:{_S1_OT150_COL}"
        ot200_col = f"'Daily Attendance'!{_S1_OT200_COL}:{_S1_OT200_COL}"

        a_ref = f"A{sum_row}"

        def _sumif_type(day_type: str) -> str:
            return f'=COUNTIFS({ref_col},{a_ref},{type_col},"{day_type}")'

        def _sumif_val(val_col: str) -> str:
            return f"=SUMIF({ref_col},{a_ref},{val_col})"

        ws2.cell(row=sum_row, column=2,
                 value=_sumif_type("Work")).font = _DATA_FONT
        ws2.cell(row=sum_row, column=3,
                 value=_sumif_type("Vacation")).font = _DATA_FONT
        ws2.cell(row=sum_row, column=4,
                 value=_sumif_type("Sick")).font = _DATA_FONT
        ws2.cell(row=sum_row, column=5,
                 value=_sumif_type("No Report")).font = _DATA_FONT
        ws2.cell(row=sum_row, column=6,
                 value=_sumif_val(pres_col)).font = _DATA_FONT
        ws2.cell(row=sum_row, column=7,
                 value=_sumif_val(pay_col)).font = _DATA_FONT
        ws2.cell(row=sum_row, column=8,
                 value=_sumif_val(ot100_col)).font = _DATA_FONT
        ws2.cell(row=sum_row, column=9,
                 value=_sumif_val(ot125_col)).font = _DATA_FONT
        ws2.cell(row=sum_row, column=10,
                 value=_sumif_val(ot150_col)).font = _DATA_FONT
        ws2.cell(row=sum_row, column=11,
                 value=_sumif_val(ot200_col)).font = _DATA_FONT

    ws2.freeze_panes = "A2"
    for col_idx in range(1, len(summary_headers) + 1):
        _auto_fit_column(ws2, col_idx)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Tkinter GUI
# ---------------------------------------------------------------------------

class AttendanceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Malam Saar Attendance – PDF to Excel")
        self.geometry("680x420")
        self.resizable(True, True)
        self._build_ui()
        self.pdf_path: str = ""

    def _build_ui(self) -> None:
        pad = {"padx": 8, "pady": 4}

        # ── Top row: file path + browse ───────────────────────────────────
        top = tk.Frame(self)
        top.pack(fill=tk.X, **pad)

        tk.Label(top, text="PDF file:").pack(side=tk.LEFT)
        self._path_var = tk.StringVar()
        tk.Entry(top, textvariable=self._path_var, width=55,
                 state="readonly").pack(side=tk.LEFT, padx=4)
        tk.Button(top, text="Browse…",
                  command=self._browse).pack(side=tk.LEFT)

        # ── Export button + progress ───────────────────────────────────────
        mid = tk.Frame(self)
        mid.pack(fill=tk.X, **pad)

        self._export_btn = tk.Button(
            mid, text="Export to Excel",
            command=self._start_export,
            bg="#1565C0", fg="white",
            font=("Arial", 10, "bold"),
            padx=12, pady=3,
        )
        self._export_btn.pack(side=tk.LEFT)

        self._progress = ttk.Progressbar(mid, mode="indeterminate", length=300)
        self._progress.pack(side=tk.LEFT, padx=8)

        self._status_var = tk.StringVar(value="Ready.")
        tk.Label(mid, textvariable=self._status_var,
                 anchor="w").pack(side=tk.LEFT, fill=tk.X, expand=True)

        # ── Log text area ─────────────────────────────────────────────────
        log_frame = tk.Frame(self)
        log_frame.pack(fill=tk.BOTH, expand=True, **pad)

        tk.Label(log_frame, text="Log:", anchor="w").pack(fill=tk.X)
        self._log_box = scrolledtext.ScrolledText(
            log_frame, height=18, font=("Courier", 9),
            state=tk.DISABLED,
        )
        self._log_box.pack(fill=tk.BOTH, expand=True)

    # ------------------------------------------------------------------
    def _browse(self) -> None:
        path = filedialog.askopenfilename(
            title="Select attendance PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if path:
            self.pdf_path = path
            self._path_var.set(path)
            self._append_log(f"Selected: {path}")

    def _start_export(self) -> None:
        if not self.pdf_path:
            self._status_var.set("No PDF selected.")
            return
        if not os.path.isfile(self.pdf_path):
            self._status_var.set("File not found.")
            return

        self._export_btn.config(state=tk.DISABLED)
        self._progress.start(10)
        self._status_var.set("Processing…")
        _log_messages.clear()

        threading.Thread(
            target=self._export_worker, daemon=True
        ).start()
        self._poll_logs()

    def _poll_logs(self) -> None:
        """Drain _log_messages into the log box every 100 ms."""
        while _log_messages:
            self._append_log(_log_messages.pop(0))
        self.after(100, self._poll_logs)

    def _export_worker(self) -> None:
        pdf = self.pdf_path
        base = os.path.splitext(pdf)[0]
        output = base + "_attendance.xlsx"
        try:
            records = process_pdf(pdf)
            if not records:
                self.after(0, self._on_error, "No attendance rows found.")
                return
            create_excel(records, output)
            self.after(0, self._on_success, output, len(records))
        except (FileNotFoundError, PermissionError, OSError) as exc:
            self.after(0, self._on_error, f"File error: {exc}")
        except (ValueError, RuntimeError, AttributeError, KeyError) as exc:
            self.after(0, self._on_error, f"{type(exc).__name__}: {exc}")

    def _on_success(self, output_path: str, count: int) -> None:
        self._progress.stop()
        self._export_btn.config(state=tk.NORMAL)
        self._status_var.set(
            f"Done – {count} rows → {os.path.basename(output_path)}"
        )
        self._append_log(f"Saved: {output_path}")

    def _on_error(self, message: str) -> None:
        self._progress.stop()
        self._export_btn.config(state=tk.NORMAL)
        self._status_var.set(f"Error: {message}")
        self._append_log(f"ERROR: {message}")

    def _append_log(self, text: str) -> None:
        self._log_box.config(state=tk.NORMAL)
        self._log_box.insert(tk.END, text + "\n")
        self._log_box.see(tk.END)
        self._log_box.config(state=tk.DISABLED)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    app = AttendanceApp()
    app.mainloop()
