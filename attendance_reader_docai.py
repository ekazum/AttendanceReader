"""
AttendanceReader (Document AI) – Malam Sachar Plus large-PDF to Excel converter.

Designed for PDFs larger than 100 pages where synchronous Document AI processing
would fail.  Uses Google Cloud Document AI **Batch Processing**, GCS for
intermediate storage, and the Document AI Toolbox library for convenient result
parsing.

Configuration is read from a local `.env` file (never committed) or the
process environment.  Required variables:

    PROJECT_ID                    – GCP project ID
    LOCATION                      – Processor region, e.g. "us" or "eu"
    PROCESSOR_ID                  – Document AI processor ID
    GCS_BUCKET_NAME               – GCS bucket for PDF upload and output
    GOOGLE_APPLICATION_CREDENTIALS – Path to the GCP service-account JSON key

Build command (PyInstaller – generates a standalone .exe):
    pyinstaller --onefile --windowed --name AttendanceReaderDocAI \\
        --exclude-module pdfplumber \\
        attendance_reader_docai.py

Note: keep .env and the credentials JSON *outside* the bundled directory;
they are loaded at runtime from the working directory.
"""

import os
import re
import threading
import uuid
import tkinter as tk
from tkinter import filedialog, ttk

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from google.cloud import documentai_v1 as documentai
from google.cloud import storage
from google.cloud.documentai_toolbox import document as docai_toolbox

# ---------------------------------------------------------------------------
# Load external configuration (.env takes priority, then env vars)
# ---------------------------------------------------------------------------

load_dotenv()  # loads .env from CWD or any parent directory


def _load_config() -> dict:
    """Return a dict of required configuration values.

    Reads from environment variables (populated by load_dotenv above, or set
    directly in the OS environment).  Raises RuntimeError with a descriptive
    message if any required variable is missing.
    """
    required = [
        "PROJECT_ID",
        "LOCATION",
        "PROCESSOR_ID",
        "GCS_BUCKET_NAME",
        "GOOGLE_APPLICATION_CREDENTIALS",
    ]
    config = {}
    missing = []
    for key in required:
        value = os.environ.get(key, "").strip()
        if not value:
            missing.append(key)
        else:
            config[key] = value

    if missing:
        raise RuntimeError(
            "חסרות הגדרות תצורה הנדרשות בקובץ .env:\n" + ", ".join(missing)
        )

    # Ensure the credentials file exists so the error is caught early.
    creds_path = config["GOOGLE_APPLICATION_CREDENTIALS"]
    if not os.path.isfile(creds_path):
        raise RuntimeError(
            f"קובץ אישורי Google Cloud לא נמצא:\n{creds_path}"
        )

    # Point the GCP client libraries at the correct credentials file.
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = creds_path
    return config


# ---------------------------------------------------------------------------
# Hebrew constants
# ---------------------------------------------------------------------------

HEBREW_DAYS = {
    "א": "ראשון",
    "ב": "שני",
    "ג": "שלישי",
    "ד": "רביעי",
    "ה": "חמישי",
    "ו": "שישי",
    "ש": "שבת",
}

DATE_RE = re.compile(r"^\d{2}/\d{2}/(\d{2}|\d{4})$")
TIME_RE = re.compile(r"^\d{2}:\d{2}(:\d{2})?$")
_HEBREW_RE = re.compile(r"[\u05d0-\u05ea]")


def fix_hebrew_text(text: str) -> str:
    """Reverse *text* if it contains Hebrew characters and is not a time or date."""
    if TIME_RE.match(text) or DATE_RE.match(text):
        return text
    if _HEBREW_RE.search(text):
        return text[::-1]
    return text


# ---------------------------------------------------------------------------
# GCS helpers
# ---------------------------------------------------------------------------

def upload_to_gcs(local_path: str, bucket_name: str, blob_name: str) -> str:
    """Upload *local_path* to GCS and return the gs:// URI."""
    client = storage.Client()
    bucket = client.bucket(bucket_name)
    blob = bucket.blob(blob_name)
    blob.upload_from_filename(local_path)
    return f"gs://{bucket_name}/{blob_name}"


def delete_from_gcs(bucket_name: str, blob_name: str) -> None:
    """Delete a GCS object; silently ignore errors (best-effort cleanup)."""
    try:
        client = storage.Client()
        bucket = client.bucket(bucket_name)
        blob = bucket.blob(blob_name)
        blob.delete()
    except Exception as exc:  # noqa: BLE001 – cleanup failure is non-fatal
        # Swallow silently; the file will eventually be cleaned up by bucket
        # lifecycle rules.  Caller can inspect exc for debugging if needed.
        _ = exc


# ---------------------------------------------------------------------------
# Document AI batch processing
# ---------------------------------------------------------------------------

def run_batch_process(
    pdf_gcs_uri: str,
    output_gcs_uri_prefix: str,
    config: dict,
    timeout: int = 600,
):
    """Submit a Document AI batch process request and wait for completion.

    Returns the long-running operation object after it finishes.
    """
    opts = {"api_endpoint": f"{config['LOCATION']}-documentai.googleapis.com"}
    client = documentai.DocumentProcessorServiceClient(
        client_options=opts
    )

    processor_name = client.processor_path(
        config["PROJECT_ID"],
        config["LOCATION"],
        config["PROCESSOR_ID"],
    )

    gcs_document = documentai.GcsDocument(
        gcs_uri=pdf_gcs_uri,
        mime_type="application/pdf",
    )
    input_config = documentai.BatchDocumentsInputConfig(
        gcs_documents=documentai.GcsDocuments(documents=[gcs_document])
    )
    output_config = documentai.DocumentOutputConfig(
        gcs_output_config=documentai.DocumentOutputConfig.GcsOutputConfig(
            gcs_uri=output_gcs_uri_prefix,
        )
    )

    request = documentai.BatchProcessRequest(
        name=processor_name,
        input_documents=input_config,
        document_output_config=output_config,
    )

    operation = client.batch_process_documents(request=request)
    operation.result(timeout=timeout)
    return operation


# ---------------------------------------------------------------------------
# Document parsing via Document AI Toolbox
# ---------------------------------------------------------------------------

def _tokens_from_page(page) -> list[str]:
    """Extract all text tokens from a Document AI Toolbox page object.

    Returns a flat list of corrected (RTL-fixed) text strings.
    """
    tokens = []
    for block in getattr(page, "blocks", []):
        for paragraph in getattr(block, "paragraphs", []):
            for token in getattr(paragraph, "tokens", []):
                text = getattr(token, "text", "").strip()
                if text:
                    tokens.append(fix_hebrew_text(text))
    return tokens


def _parse_attendance_from_tokens(tokens: list[str]) -> list[dict]:
    """Parse attendance records from a flat list of page tokens.

    The Malam Sachar Plus layout is a horizontal grid; we rely on positional
    patterns in the token stream to identify rows rather than exact coordinates
    (coordinate data is not always reliable across Document AI shards).
    """
    records = []
    dates: list[str] = [t for t in tokens if DATE_RE.match(t)]
    if not dates:
        return []

    # Attempt to align times to dates by finding labelled sequences.
    # We look for the labels and take the next N tokens (where N == len(dates))
    # immediately following the label as the values for each date column.
    def _extract_row_after_label(tokens: list[str], label_substr: str) -> list[str]:
        """Return up to len(dates) time tokens that follow a label token."""
        for i, t in enumerate(tokens):
            if label_substr in t:
                # Collect subsequent time tokens
                values = []
                for j in range(i + 1, len(tokens)):
                    if TIME_RE.match(tokens[j]):
                        values.append(tokens[j])
                        if len(values) == len(dates):
                            break
                    elif DATE_RE.match(tokens[j]):
                        # Hit a new date row – stop
                        break
                return values
        return []

    def _extract_row_day_letters(tokens: list[str]) -> list[str]:
        letters = [t for t in tokens if t in HEBREW_DAYS]
        return letters[: len(dates)]

    def _extract_row_free_text(tokens: list[str], label_substr: str) -> list[str]:
        for i, t in enumerate(tokens):
            if label_substr in t:
                values = []
                for j in range(i + 1, len(tokens)):
                    tok = tokens[j]
                    if DATE_RE.match(tok) or TIME_RE.match(tok) or tok in HEBREW_DAYS:
                        break
                    if _HEBREW_RE.search(tok) and len(tok) > 1:
                        values.append(tok)
                        if len(values) == len(dates):
                            break
                return values
        return []

    day_letters = _extract_row_day_letters(tokens)
    entry_times = _extract_row_after_label(tokens, "כניסה")
    exit_times = _extract_row_after_label(tokens, "יציאה")
    total_times = _extract_row_after_label(tokens, "נוכח")
    day_types = _extract_row_free_text(tokens, "סוג")

    for idx, date in enumerate(dates):
        records.append(
            {
                "date": date,
                "day": HEBREW_DAYS.get(
                    day_letters[idx] if idx < len(day_letters) else "", ""
                ),
                "entry": entry_times[idx] if idx < len(entry_times) else "",
                "exit": exit_times[idx] if idx < len(exit_times) else "",
                "total": total_times[idx] if idx < len(total_times) else "",
                "day_type": day_types[idx] if idx < len(day_types) else "",
            }
        )
    return records


def parse_docai_results(operation, config: dict) -> list[dict]:
    """Use Document AI Toolbox to retrieve and parse batch results.

    Returns a list of attendance record dicts.
    """
    wrapped_documents = docai_toolbox.Document.from_batch_process_operation(
        location=config["LOCATION"],
        operation_name=operation.operation.name,
    )

    all_records: list[dict] = []
    for doc in wrapped_documents:
        for page in doc.pages:
            tokens = _tokens_from_page(page)
            all_records.extend(_parse_attendance_from_tokens(tokens))

    return all_records


# ---------------------------------------------------------------------------
# Excel output (reuses same structure as attendance_reader.py)
# ---------------------------------------------------------------------------

HEADERS = [
    "תאריך",
    "יום בשבוע",
    "שעת כניסה",
    "שעת יציאה",
    'סה"כ שעות ליום',
    "סוג יום",
]
COLUMN_WIDTHS = [12, 15, 14, 14, 18, 14]


def create_excel(records: list, output_path: str) -> None:
    """Write *records* to an Excel file at *output_path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "נוכחות"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    data_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, record in enumerate(records, 2):
        values = [
            record.get("date", ""),
            record.get("day", ""),
            record.get("entry", ""),
            record.get("exit", ""),
            record.get("total", ""),
            record.get("day_type", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = data_font
            cell.alignment = center

    for col, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.auto_filter.ref = f"A1:F{len(records) + 1}"
    ws.freeze_panes = "A2"

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Tkinter GUI
# ---------------------------------------------------------------------------

class AttendanceDocAIApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ממיר נוכחות (Document AI) – PDF לאקסל")
        self.geometry("600x260")
        self.resizable(False, False)
        self._configure_encoding()

        self.pdf_path = tk.StringVar()
        self._build_ui()

    def _configure_encoding(self):
        self.option_add("*font", "Arial 10")
        try:
            self.tk.call("encoding", "system", "utf-8")
        except tk.TclError:
            pass

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        # ── PDF selection ──────────────────────────────────────────────────
        row1 = tk.Frame(self)
        row1.pack(fill=tk.X, **pad)

        tk.Button(row1, text="בחר קובץ PDF", command=self._browse_pdf,
                  width=14).pack(side=tk.RIGHT)
        tk.Entry(row1, textvariable=self.pdf_path, width=48,
                 state="readonly").pack(side=tk.RIGHT, padx=(0, 6))

        # ── Convert button ─────────────────────────────────────────────────
        row2 = tk.Frame(self)
        row2.pack(fill=tk.X, **pad)

        self.convert_btn = tk.Button(
            row2, text="עבד עם Document AI  ▶",
            command=self._start_conversion,
            bg="#1565C0", fg="white",
            font=("Arial", 11, "bold"),
            padx=14, pady=4,
        )
        self.convert_btn.pack(side=tk.RIGHT)

        # ── Status label ───────────────────────────────────────────────────
        row3 = tk.Frame(self)
        row3.pack(fill=tk.X, **pad)

        self.status_var = tk.StringVar(
            value="נא לבחור קובץ PDF ולחץ 'עבד עם Document AI'"
        )
        tk.Label(
            row3, textvariable=self.status_var,
            wraplength=580, justify="right", anchor="e",
        ).pack(fill=tk.X)

        # ── Progress bar ───────────────────────────────────────────────────
        self.progress = ttk.Progressbar(self, mode="indeterminate")
        self.progress.pack(fill=tk.X, padx=10, pady=(0, 8))

        # ── Configuration hint ─────────────────────────────────────────────
        hint = tk.Label(
            self,
            text="הגדרות נטענות מקובץ .env בתיקיית ההרצה",
            font=("Arial", 8),
            fg="#666666",
        )
        hint.pack(anchor="e", padx=10, pady=(0, 4))

    # ------------------------------------------------------------------
    # Event handlers
    # ------------------------------------------------------------------

    def _browse_pdf(self):
        path = filedialog.askopenfilename(
            title="בחר קובץ נוכחות PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if path:
            self.pdf_path.set(path)
            self.status_var.set(f"נבחר: {os.path.basename(path)}")

    def _start_conversion(self):
        pdf = self.pdf_path.get()
        if not pdf:
            self.status_var.set("שגיאה: לא נבחר קובץ PDF")
            return
        if not os.path.isfile(pdf):
            self.status_var.set("שגיאה: הקובץ אינו קיים")
            return

        default_name = os.path.splitext(os.path.basename(pdf))[0] + ".xlsx"
        output = filedialog.asksaveasfilename(
            title="שמור קובץ אקסל",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_name,
        )
        if not output:
            return

        self.convert_btn.config(state=tk.DISABLED)
        self.progress.start(12)
        self.status_var.set("שולח ל-Document AI, נא להמתין (עד כמה דקות)…")

        threading.Thread(
            target=self._convert_worker, args=(pdf, output), daemon=True
        ).start()

    def _convert_worker(self, pdf_path: str, output_path: str):
        blob_name = None
        config = None
        try:
            # 1. Load configuration
            self.after(0, self.status_var.set, "טוען הגדרות תצורה…")
            config = _load_config()

            # 2. Upload PDF to GCS
            unique_prefix = uuid.uuid4().hex[:8]
            blob_name = f"attendance-uploads/{unique_prefix}/{os.path.basename(pdf_path)}"
            gcs_output_prefix = (
                f"gs://{config['GCS_BUCKET_NAME']}/attendance-output/{unique_prefix}/"
            )

            self.after(0, self.status_var.set, "מעלה קובץ PDF ל-GCS…")
            pdf_gcs_uri = upload_to_gcs(
                pdf_path, config["GCS_BUCKET_NAME"], blob_name
            )

            # 3. Submit batch process and await completion
            self.after(0, self.status_var.set,
                       "Document AI מעבד את המסמך (עד 10 דקות)…")
            operation = run_batch_process(
                pdf_gcs_uri=pdf_gcs_uri,
                output_gcs_uri_prefix=gcs_output_prefix,
                config=config,
            )

            # 4. Parse results via Toolbox
            self.after(0, self.status_var.set, "מנתח תוצאות…")
            records = parse_docai_results(operation, config)

            # 5. Cleanup – delete uploaded PDF from GCS
            delete_from_gcs(config["GCS_BUCKET_NAME"], blob_name)
            blob_name = None  # already deleted

            if not records:
                self.after(0, self._on_error, "לא נמצאו נתוני נוכחות בקובץ")
                return

            # 6. Write Excel
            create_excel(records, output_path)
            self.after(0, self._on_success, output_path, len(records))

        except (FileNotFoundError, PermissionError, OSError) as exc:
            self.after(0, self._on_error, f"שגיאת קובץ: {exc}")
        except RuntimeError as exc:
            self.after(0, self._on_error, str(exc))
        except Exception as exc:  # noqa: BLE001
            self.after(0, self._on_error, f"{type(exc).__name__}: {exc}")
        finally:
            # Best-effort cleanup if an error occurred after the upload
            if blob_name and config:
                delete_from_gcs(config["GCS_BUCKET_NAME"], blob_name)

    def _on_success(self, output_path: str, count: int):
        self.progress.stop()
        self.convert_btn.config(state=tk.NORMAL)
        self.status_var.set(
            f"✔ הצלחה! {count} ימים יוצאו → {os.path.basename(output_path)}"
        )

    def _on_error(self, message: str):
        self.progress.stop()
        self.convert_btn.config(state=tk.NORMAL)
        self.status_var.set(f"✘ שגיאה: {message}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    app = AttendanceDocAIApp()
    app.mainloop()
